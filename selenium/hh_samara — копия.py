# ◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾
# Парсинг вакансий https://zarplata.ru/
# ◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾

release = '(1.01)'

import os, re, sqlite3, datetime, locale

from openpyxl import Workbook


from time import sleep

from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.common.exceptions import *

from lxml import etree

# ◾◾◾◾◾◾◾◾◾◾◾◾ ГЛАВНЫЙ

def main():


    now = datetime.datetime.now()
    date_now = now.strftime("%d.%m")
    
    # ◾◾◾ создание БД SQLite
    
    file_db = 'db/hh_samara.db'
    con = sqlite3.connect(file_db)
    con.row_factory = sqlite3.Row                                      # для доступа к полям по имени
    query = con.cursor()
    # query.execute("CREATE TABLE IF NOT EXISTS vacant "
            # "(rec_id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL ,"
            # "marker TEXT DEFAULT '',"
            # "link TEXT type UNIQUE,"
            # "vacancy TEXT DEFAULT '',"
            # "company TEXT DEFAULT '',"
            # "phone TEXT DEFAULT '',"
            # "face TEXT DEFAULT '', "
            # "date TEXT DEFAULT '');")


    # # ◾◾◾ запускаем драйвер Chrome
    
    # path_to_chromedriver = 'chromedriver' # где лежит WebDriver
    # driver = webdriver.Chrome(executable_path = path_to_chromedriver) # запуск WebDriver
    

    # # ◾◾◾ собираем ссылки на вакансии 50 страниц по 20 вакансий = 1000 вакансий

    # for i in range(0,50,1): # до 50
    
        # url = "https://togliatti.hh.ru/search/vacancy?L_is_autosearch=false&area=78&clusters=true&enable_snippets=true&page={0}".format(i)
        # # print(url)
        # driver.get(url)
        # sleep(2)
        # links = driver.find_elements_by_xpath('//a[@class ="bloko-link HH-LinkModifier"]')
        
        # for link in links:  
        
            # # print (link.get_attribute('href'))

            # # ◾◾◾ записать в базу данных ссылки
            # try:
                # query.execute("INSERT OR IGNORE INTO vacant (link) VALUES (?) ",
                # (link.get_attribute('href'),))

            # except sqlite3.Error as e:
                # print (e.args[0])

            # con.commit()
        
    # # ◾◾◾ читаем линки из БД. только не заполненные
    
    # query.execute("SELECT rec_id, link FROM vacant WHERE date= '';")
    # rec_all = query.fetchall()
    # for rec in rec_all:

        # url = rec['link']
        # print(url)
        # driver.get(url) # загрузка страницы
        # sleep(2)

        # # ◾◾◾ вакансия компания и дата размещения
        
        # try:
            # element = driver.find_element_by_css_selector(".vacancy-title > h1")
            # vacancy = element.text
        # except NoSuchElementException:
            # continue
        
        # try:
            # element = driver.find_element_by_xpath('//span[@itemprop="name"]') 
            # company = element.text
        # except NoSuchElementException:
            # continue
        
        # try:
            # element = driver.find_element_by_class_name("vacancy-creation-time")
            # date = element.text
        # except NoSuchElementException:
            # continue

        # # ◾◾◾ открыть контакты 
        
        # try:
            # button = driver.find_element_by_class_name("HH-VacancyContacts-Switcher") 
            # button.click()
            # sleep(1)
            
            # # ◾◾◾ телефон
            
            # try:
                # element = driver.find_element_by_xpath("//*[@data-qa='vacancy-contacts__phone']") 
                # phone = element.text
            # except NoSuchElementException:
                # phone = ''
                
            # # ◾◾◾ контактное лицо
            
            # try:
                # element = driver.find_element_by_xpath("//*[@data-qa='vacancy-contacts__fio']") 
                # face = element.text
            # except NoSuchElementException:
                # face = ''
                
        # except NoSuchElementException:
            # sleep(1)
            # phone = ''
            # face = ''
            # # print('нет кнопки')

        # # print (vacancy)
        # # print (company)
        # # print (phone)
        # # print (face)
        # # print (date)
        
        # query.execute(
                # "UPDATE vacant SET marker = ?,vacancy = ?,company = ?,phone = ?,face = ?,date = ? "
                # "WHERE rec_id = ?",
                # (date_now, vacancy, company, phone, face, date, rec['rec_id'])
            # )
        # con.commit()


    # driver.quit()

    query.execute("SELECT * FROM vacant")
    rec_all = query.fetchall()

    # ◾◾◾ запись в файл excel

    wb = Workbook()
    ws = wb.active
    ws.title = date_now

    i = 1
    for rec in rec_all:
    
        ws.cell(row=i, column=1).value = rec['link']
        ws.cell(row=i, column=2).value = rec['vacancy']
        ws.cell(row=i, column=3).value = rec['company']
        ws.cell(row=i, column=4).value = rec['phone']
        ws.cell(row=i, column=5).value = rec['face']
        ws.cell(row=i, column=6).value = rec['date']
        i += 1
    
    file_xlsx = os.getenv('HOME') + '\\Documents\\hh_samara.xlsx'
    wb.save(file_xlsx)

    con.close()
    


if __name__ == '__main__':                                                      # ОСНОВНОЙ ЦИКЛ
    main()
    
   
    

