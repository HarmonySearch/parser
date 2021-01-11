# ◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾
# Парсинг вакансий https://zarplata.ru/
# ◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾◾

release = '(1.01)'

import os, re, sqlite3, datetime, locale

from openpyxl import Workbook


from time import sleep

from selenium.webdriver.common.keys import Keys
from selenium import webdriver
from selenium.common.exceptions import *

from lxml import etree

# ◾◾◾◾◾◾◾◾◾◾◾◾ ГЛАВНЫЙ

def main():


    now = datetime.datetime.now()
    date_now = now.strftime("%d.%m")
    
    # ◾◾◾ создание БД SQLite
    
    file_db = 'db/hh_samara.db'
    con = sqlite3.connect(file_db)
    con.row_factory = sqlite3.Row                                      # для доступа к полям по имени
    query = con.cursor()
    query.execute("CREATE TABLE IF NOT EXISTS vacant "
            "(rec_id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL ,"
            "marker TEXT DEFAULT '',"
            "link TEXT type UNIQUE,"
            "vacancy TEXT DEFAULT '',"
            "company TEXT DEFAULT '',"
            "phone TEXT DEFAULT '',"
            "face TEXT DEFAULT '', "
            "date TEXT DEFAULT '');")


    
    query.execute("SELECT * FROM vacant WHERE marker= '{}';".format(date_now))
    rec_all = query.fetchall()

    # ◾◾◾ запись в файл excel

    wb = Workbook()
    ws = wb.active
    ws.title = date_now

    i = 1
    for rec in rec_all:
    
        ws.cell(row=i, column=1).value = rec['link']
        ws.cell(row=i, column=2).value = rec['vacancy']
        ws.cell(row=i, column=3).value = rec['company']
        ws.cell(row=i, column=4).value = rec['phone']
        ws.cell(row=i, column=5).value = rec['face']
        ws.cell(row=i, column=6).value = rec['date']
        i += 1
    
    file_xlsx = os.getenv('HOME') + '\\Documents\\hh_samara.xlsx'
    wb.save(file_xlsx)

    con.close()
    


if __name__ == '__main__':                                                      # ОСНОВНОЙ ЦИКЛ
    main()
    
   
    

