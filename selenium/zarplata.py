# ---------------------------------------------------------------------
# Парсинг вакансий https://zarplata.ru/
# ---------------------------------------------------------------------

import os
import sys
import locale
import datetime
import sqlite3
import re
from lxml import etree
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import *
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from time import sleep

from openpyxl import Workbook


release = '(1.01)'


# import pickle


def main():

    login():
    find_link():
    find_info():
    write_excel():



login():



find_link():

    now = datetime.datetime.now()

    locale.setlocale(locale.LC_ALL, 'ru_RU.UTF-8')

    date_now = now.strftime("%d %b")

    # создание папки
    db_dir = 'dbase'
    if not os.path.exists(db_dir):
        os.mkdir(db_dir)

    # создание базы данных SQLite
    db_sqlite3 = db_dir + '/db_sqlite3.db'
    con = sqlite3.connect(db_sqlite3)
    con.row_factory = sqlite3.Row                                               # для доступа к полям по имени
    query = con.cursor()
    query.execute("CREATE TABLE IF NOT EXISTS vacant "
                  "(rec_id INTEGER PRIMARY KEY AUTOINCREMENT NOT NULL ,"
                  "link TEXT DEFAULT '',"
                  "vacancy TEXT DEFAULT '',"
                  "company TEXT DEFAULT '',"
                  "phone TEXT DEFAULT '',"
                  "face TEXT DEFAULT '', "
                  "date TEXT DEFAULT '');")
    # query.close()

    path_to_chromedriver = 'driver/chromedriver'                                # адрес WebDriver
    
    # для запуска без окна хрома
    chrome_options = Options()
    chrome_options.add_argument('--headless')
    chrome_options.add_argument('--no-sandbox')
    chrome_options.add_argument('--window-size=1920,1080')
    chrome_options.add_argument('--proxy-server="direct://"')
    chrome_options.add_argument('--proxy-bypass-list=*')
    driver = webdriver.Chrome(executable_path=path_to_chromedriver, options=chrome_options) 

    # загрузка страницы
    url = 'https://tolyatti.zarplata.ru/'
    driver.get(url)
    sleep(3)

    # Залогонимся на сайте, а то не видно будет ничего полезного
    # Находим кнопку Вход. У кнопки есть атрибут  href, по которому происходит переход на форму регистрации.
    # Грузим страницу входа (get) и ищем на странице поле логина (find_element_by_class_name).
    # Вводим текст (send_keys() в поле логина. Ищем поле ввода пароля (input_1Um_6).
    # Вводим туда пароль (send_keys(). Нажимаем кнопку отправить (submit).
    # Ждем секунды 3, пока куки не прогрузятся гарантированно.

    new_window_url = driver.find_element_by_link_text(
        "Вход").get_attribute("href")
    driver.get(new_window_url)
    sleep(3)
    element = driver.find_element_by_class_name("input")
    element.send_keys("rabotarutlt@gmail.com")
    element = driver.find_element_by_class_name("input_1Um_6")
    element.send_keys("m3X8P49T9f92DMWkJf5#")
    element.submit()
    sleep(3)
    
    # --- читаем линки вакансий с сата ----------
    
    # первая страница вакансий

    # url = "https://tolyatti.zarplata.ru/vacancy"
    # driver.get(url)
    # while True:
    #     # пока не появится кнопка 'Дальше'
    #     try:
    #         element = WebDriverWait(driver, 10).until(
    #             EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Дальше'))
    #         )
    #     except:
    #         print("Кнопка не зпгрузилась")
    #         sys.exit()
    #     finally:
    #         print("загрузка прошла")

    #     print(len(driver.find_elements_by_class_name('vacancy-title_1sIep')))

    #     # читаем все ссылки на описание вакансий на странице и грузим в базу данных

    #     for vacant in driver.find_elements_by_class_name('vacancy-title_1sIep'):
    #             link = vacant.get_attribute('href')
    #             print(link)

    #             # записать в базу данных ссылки
    #             try:
    #                 query.execute("INSERT INTO vacant (link) VALUES (?);", (link,))

    #             except sqlite3.Error as e:
    #                 print(e.args[0])
    #                 # logging.critical('btn_collect_ref_click: {} {}'.format( e.args[0],sql))

    #             con.commit()


    #     # если link (кнопка с текстом "Дальше..." имеет класс disabled то выход

    #     btn_next = driver.find_element_by_partial_link_text('Дальше')
    #     vclass = btn_next.get_attribute("class")
    #     print('+')
    #     if vclass.find('disabled') == -1:
    #     # нажимаем кнопку Дальше
    #         btn_next.click()
    #     else:
    #         break

    # print(vclass,vclass.find('disabled'))



    # driver.quit()
    # sys.exit()


    def find_info():

    # --- заполняем информацией линки из базы данных ----------

    # читаем линки из БД. только не заполненные

    query.execute("SELECT rec_id, link FROM vacant;")
    rec_all = query.fetchall()
    for rec in rec_all:

        # загрузка страницы
        url = rec['link']
        driver.get(url)         

        try:
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.PARTIAL_LINK_TEXT, 'Пожаловаться'))
            )
        except:
            print("Страница  не загрузилась")
            sys.exit()
        finally:
            print("загрузка прошла")


        # sleep(3)

        # название вакансии единственный тег h1. Есть обязательно
        element = driver.find_element_by_tag_name('h1') 
        vacancy = element.text

        print ('вакансия-'+vacancy)

        # нужно поднятся на два ула вверх и упасть до второго линка. первый линк звезда.
        parent = element.find_element_by_xpath("../..")
        element = parent.find_element_by_tag_name('div h2 a')
        company = element.text

        # группа Контактная информация имеет внятное имя.
        # Узел необязательный - если нет, то следующая вакансия
        try:
            contact_info = driver.find_element_by_name('vacancy-contacts')
        except NoSuchElementException:
            continue

        # контактное лицо
        try:
            element = contact_info.find_element_by_xpath("./div/div/div")
            face = element.text
        except NoSuchElementException:
            face = 'нет данных'

        # контактный телефон с кнопкой развертывания
        try:
            btn = contact_info.find_element_by_css_selector("button")
            element_phone = btn.find_element_by_xpath('..')
            btn.click()
            phone = element_phone.text
        except NoSuchElementException:
            phone = 'нет данных'
        
        date = 'второй заход'

        print(vacancy) # название вакансии
        print(company) #  компания
        print(phone) #  телефон
        print(face) # контактное лицо
        print(date) # дата возникновения?

        sql = ("UPDATE vacant SET "
               "vacancy = '{}',"
               "company = '{}',"
               "phone = '{}',"
               "face = '{}',"
               "date = '{}' "
               "WHERE rec_id = {}"
               .format(vacancy, company, phone, face, date, rec['rec_id']))
        query.execute(sql)
        con.commit()


# Записываем базу в екселевскую таблицу

def write_excel():

    query.execute("SELECT * FROM vacant WHERE marker= '{}';".format(date_now))
    rec_all = query.fetchall()

    # ◾◾◾ запись в файл excel

    wb = Workbook()
    ws = wb.active
    ws.title = date_now

    i = 1
    for rec in rec_all:
    
        ws.cell(row=i, column=1).value = rec['link']
        ws.cell(row=i, column=2).value = rec['vacancy']
        ws.cell(row=i, column=3).value = rec['company']
        ws.cell(row=i, column=4).value = rec['phone']
        ws.cell(row=i, column=5).value = rec['face']
        ws.cell(row=i, column=6).value = rec['date']
        i += 1
    
    file_xlsx = 'hh_samara.xlsx'
    wb.save(file_xlsx)

    con.close()
    


if __name__ == '__main__':                                                      # ОСНОВНОЙ ЦИКЛ
    main()
